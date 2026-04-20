"""
matrixify_erp_filler.py
=======================
Reads a Matrixify Orders export (.xlsx), looks up GP ERP numbers from
RRPRead.[dbo].[CustomerShippingDetail], populates:
  - Metafield: custom.erp_master_order_number [single_line_text_field]
  - Metafield: custom.erp_invoice_number [single_line_text_field]
then writes one output .xlsx per brand for Matrixify import.

Join key: Shopify order Name (e.g. '#YUK4699') → GP D.[PO#] (e.g. 'YUK4699')
Brand is detected from leading alpha characters in the PO# (YUK, RWW, ZUM, etc.)
"""

import os
import sys
import re
import argparse
import datetime
import traceback

import pandas as pd
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Import credentials ─────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

try:
    from config import DB_CONFIG
except ImportError:
    print("ERROR: config.py not found. Make sure config.py is in the same folder as this script.")
    sys.exit(1)

# ── Column names in the Matrixify export ──────────────────────────────────────
COL_ID       = "ID"
COL_COMMAND  = "Command"
COL_NAME     = "Name"
COL_MASTER   = "Metafield: custom.erp_master_order_number [single_line_text_field]"
COL_INVOICE  = "Metafield: custom.erp_invoice_number [single_line_text_field]"

# Output columns for the Matrixify import file
OUTPUT_COLS = [COL_ID, COL_COMMAND, COL_NAME, COL_MASTER, COL_INVOICE]

# ── SQL Query ──────────────────────────────────────────────────────────────────
# Fetches one row per PO# (takes the first INVOICE# and MASTER# if multiple exist)
SQL_QUERY = """
SELECT
    D.[PO#],
    MIN(D.[MASTER#]) AS [MASTER#],
    MIN(D.[INVOICE#]) AS [INVOICE#]
FROM [RRPRead].[dbo].[CustomerShippingDetail] D
WHERE D.[PO#] IN ({placeholders})
GROUP BY D.[PO#]
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def get_best_odbc_driver():
    preferred = [
        "ODBC Driver 18 for SQL Server",
        "ODBC Driver 17 for SQL Server",
        "ODBC Driver 13 for SQL Server",
        "SQL Server Native Client 11.0",
        "SQL Server",
    ]
    available = pyodbc.drivers()
    for driver in preferred:
        if driver in available:
            return driver
    raise RuntimeError(
        f"No suitable SQL Server ODBC driver found.\n"
        f"Available: {available}\n"
        f"Download from: https://aka.ms/downloadmsodbcsql"
    )


def get_db_connection():
    """Connect to SQL Server using Windows Integrated Security."""
    driver = get_best_odbc_driver()
    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={DB_CONFIG['server']};"
        f"DATABASE={DB_CONFIG['database']};"
        f"Trusted_Connection=yes;"
        f"Encrypt=yes;"
        f"TrustServerCertificate=yes;"
        f"Connection Timeout=30;"
    )
    return pyodbc.connect(conn_str)


def extract_po(name):
    """Strip leading '#' from Shopify order name to get the GP PO#."""
    if not name:
        return None
    return str(name).lstrip("#").strip()


def detect_brand(po):
    """Return the leading alpha prefix of a PO# as the brand label."""
    if not po:
        return "NOPFX"
    m = re.match(r"^([A-Za-z]+)", po)
    return m.group(1).upper() if m else "NOPFX"


def fetch_erp_data(po_list):
    """Query GP for MASTER# and INVOICE# for the given PO# values."""
    if not po_list:
        return pd.DataFrame(columns=["PO#", "MASTER#", "INVOICE#"])

    print(f"  Connecting to {DB_CONFIG['server']} / {DB_CONFIG['database']}...")
    conn = get_db_connection()

    # Query in batches of 1000 (SQL Server IN clause limit)
    results = []
    batch_size = 1000
    for i in range(0, len(po_list), batch_size):
        batch = po_list[i : i + batch_size]
        placeholders = ",".join(["?" for _ in batch])
        query = SQL_QUERY.format(placeholders=placeholders)
        df_batch = pd.read_sql(query, conn, params=batch)
        results.append(df_batch)
        print(f"  Fetched batch {i // batch_size + 1}: {len(df_batch)} matches")

    conn.close()

    if results:
        return pd.concat(results, ignore_index=True)
    return pd.DataFrame(columns=["PO#", "MASTER#", "INVOICE#"])


def style_header_row(ws):
    """Apply header styling to the first row of a worksheet."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def write_brand_file(df_brand, brand, output_dir, date_str):
    """Write a single brand's data to an xlsx file formatted for Matrixify import."""
    filename = f"{brand}_erp_{date_str}.xlsx"
    filepath = os.path.join(output_dir, filename)
    cols = list(df_brand.columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Write header
    ws.append(cols)

    # Write data rows
    for _, row in df_brand.iterrows():
        ws.append([row.get(col, "") for col in cols])

    # Style header
    style_header_row(ws)

    # Auto-fit columns
    for col_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)), 12)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


def run(input_path, output_dir=None):
    print(f"\n{'='*60}")
    print(f"  Matrixify ERP Filler")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    # ── Step 1: Read the Matrixify export ────────────────────────────────────
    print(f"[1/5] Reading Matrixify export: {input_path}")
    try:
        # Try "Orders" sheet first, fall back to first sheet
        import openpyxl as _opx
        wb_check = _opx.load_workbook(input_path, read_only=True)
        sheet_names = wb_check.sheetnames
        wb_check.close()
        sheet = "Orders" if "Orders" in sheet_names else sheet_names[0]
        print(f"  Using sheet: '{sheet}' (available: {sheet_names})")
        df = pd.read_excel(input_path, sheet_name=sheet, dtype=str)
        df = df.where(pd.notna(df), None)
    except Exception as e:
        print(f"  ERROR reading file: {e}")
        sys.exit(1)

    print(f"  Loaded {len(df)} orders.")

    # Accept either 'Name' or 'Number' as the order name column
    global COL_NAME
    if COL_NAME not in df.columns:
        for alt in ["Number", "Name", "Order Name", "Order Number"]:
            if alt in df.columns:
                COL_NAME = alt
                print(f"  Using '{COL_NAME}' as order name column")
                break
        else:
            print(f"\n  ERROR: Could not find order name column.")
            print(f"  Available columns: {list(df.columns)}")
            sys.exit(1)

    # Ensure metafield columns exist (add if missing)
    for col in [COL_MASTER, COL_INVOICE]:
        if col not in df.columns:
            df[col] = None

    # Update OUTPUT_COLS to use the actual name column
    output_cols = [COL_ID, COL_COMMAND, COL_NAME, COL_MASTER, COL_INVOICE]

    # ── Step 2: Extract PO#s, detect brand, skip numeric-only orders ─────────
    print("[2/5] Extracting PO# and detecting brand...")
    df["_PO#"]    = df[COL_NAME].apply(extract_po)
    df["_Brand"]  = df["_PO#"].apply(detect_brand)
    df[COL_COMMAND] = "UPDATE"

    total_before = len(df)
    df = df[df["_Brand"] != "NOPFX"].copy()
    skipped = total_before - len(df)
    if skipped:
        print(f"  Skipped {skipped} orders with no brand prefix (numeric-only names)")

    if df.empty:
        print(f"\n  WARNING: No branded orders found. All {total_before} orders had numeric-only names.")
        sys.exit(1)

    brand_counts = df["_Brand"].value_counts().to_dict()
    print(f"  Brands to export: {dict(sorted(brand_counts.items()))}")

    po_list = [po for po in df["_PO#"].dropna().unique().tolist() if po]
    print(f"  Unique PO#s to look up: {len(po_list)}")

    # ── Step 3: Query GP ──────────────────────────────────────────────────────
    print("[3/5] Querying GP (RRPRead.dbo.CustomerShippingDetail)...")
    try:
        df_erp = fetch_erp_data(po_list)
    except Exception as e:
        print(f"  ERROR querying GP: {e}")
        traceback.print_exc()
        sys.exit(1)

    print(f"  GP returned {len(df_erp)} matched orders.")

    # Normalize PO# for join (strip whitespace)
    df_erp["PO#"] = df_erp["PO#"].astype(str).str.strip()
    df["_PO#_norm"] = df["_PO#"].astype(str).str.strip()

    # ── Step 4: Join GP data back ─────────────────────────────────────────────
    print("[4/5] Joining ERP data to Matrixify orders...")
    erp_lookup = df_erp.set_index("PO#")[["MASTER#", "INVOICE#"]].to_dict("index")

    matched = 0
    unmatched = []
    for idx, row in df.iterrows():
        po = row["_PO#_norm"]
        if po in erp_lookup:
            df.at[idx, COL_MASTER]  = str(erp_lookup[po]["MASTER#"]).strip()
            df.at[idx, COL_INVOICE] = str(erp_lookup[po]["INVOICE#"]).strip()
            matched += 1
        else:
            unmatched.append(row.get(COL_NAME, po))

    print(f"  Matched: {matched} / {len(df)} orders")
    if unmatched:
        print(f"  No GP match for {len(unmatched)} orders: {unmatched[:10]}"
              + (" ..." if len(unmatched) > 10 else ""))

    # ── Step 5: Write one output file per brand ───────────────────────────────
    print("[5/5] Writing output file(s)...")

    if output_dir is None:
        output_dir = os.path.join(SCRIPT_DIR, "matrixify_output")
    os.makedirs(output_dir, exist_ok=True)

    date_str = datetime.datetime.now().strftime("%Y%m%d")
    written_files = []

    for brand in sorted(df["_Brand"].unique()):
        df_brand = df[df["_Brand"] == brand][output_cols].copy()
        filepath = write_brand_file(df_brand, brand, output_dir, date_str)
        print(f"  [{brand}] {len(df_brand)} orders → {os.path.basename(filepath)}")
        written_files.append(filepath)

    print(f"\n✅ Done! {len(written_files)} file(s) written to: {output_dir}\n")
    for f in written_files:
        print(f"   {f}")
    print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Fill ERP order/invoice numbers into a Matrixify export, split by brand."
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="Path to the Matrixify Orders export .xlsx file. "
             "If omitted, looks for the most recent .xlsx in the script folder.",
    )
    parser.add_argument(
        "--output-dir", "-o",
        default=None,
        help="Output directory for brand files. Defaults to 'matrixify_output/' next to the script.",
    )
    args = parser.parse_args()

    # Auto-detect input file if not provided
    if args.input:
        input_path = args.input
    else:
        # Look for most recent xlsx in script dir that isn't an output file
        xlsx_files = [
            f for f in os.listdir(SCRIPT_DIR)
            if f.lower().endswith(".xlsx") and "erp_" not in f.lower()
        ]
        if not xlsx_files:
            print("ERROR: No .xlsx input file found. Pass the file path as an argument.")
            sys.exit(1)
        # Pick most recently modified
        xlsx_files.sort(key=lambda f: os.path.getmtime(os.path.join(SCRIPT_DIR, f)), reverse=True)
        input_path = os.path.join(SCRIPT_DIR, xlsx_files[0])
        print(f"Auto-detected input file: {input_path}")

    run(input_path, args.output_dir)
